def convert():
    import random
    strokaeng = random.randint(1, 242)
    import openpyxl
    wb = openpyxl.load_workbook(filename='eng.xlsx')
    # получаем активный лист
    sheet = wb.active
    cell = sheet.cell(row=strokaeng, column=1)
    print('Значение: ' + cell.value)
    strokarus = strokaeng

    cell2 = sheet.cell(row=strokarus, column=2)
    perevod = input('Перевод - ')
    print('Правильный ответ: ' + cell2.value)

    if perevod == cell2.value:
        print('правильный перевод')
    else:
        print('неправильный перевод')
while True:
    flag = input('Ещё раз? [да/нет]: ')

    if flag == 'да':
        convert()
    else:
        break
